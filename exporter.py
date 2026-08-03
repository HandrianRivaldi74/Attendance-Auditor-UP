"""
exporter.py
Ekspor hasil pemeriksaan absensi (anomali) dan hasil banding absensi vs struk
gaji ke satu file Excel (beberapa sheet).
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_FONT = Font(bold=True)
BAD_FILL = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
OK_FILL = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")


def _tulis_header(ws, kolom):
    for i, judul in enumerate(kolom, start=1):
        c = ws.cell(row=1, column=i, value=judul)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    ws.freeze_panes = "A2"


def _lebarkan_kolom(ws, lebar):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(lebar, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def ekspor_hasil(path_output, anomali_absensi=None, hasil_banding=None, ringkasan_banding=None):
    wb = Workbook()

    # Sheet 1: Anomali Absensi (4 aturan)
    ws1 = wb.active
    ws1.title = "Anomali Absensi"
    _tulis_header(ws1, ["NRP", "Nama", "Aturan", "Detail", "Tanggal"])
    _lebarkan_kolom(ws1, [12, 28, 12, 60, 10])
    if anomali_absensi:
        for r, a in enumerate(anomali_absensi, start=2):
            ws1.cell(r, 1, a.nrp)
            ws1.cell(r, 2, a.nama)
            ws1.cell(r, 3, a.aturan)
            ws1.cell(r, 4, a.detail)
            ws1.cell(r, 5, a.tanggal)
            for c in range(1, 6):
                ws1.cell(r, c).fill = BAD_FILL

    # Sheet 2: Banding Absensi vs Struk Gaji
    ws2 = wb.create_sheet("Banding Absensi vs Gaji")
    _tulis_header(ws2, ["NRP", "Nama (Absensi)", "Nama (Struk Gaji)", "Status", "Detail Perbedaan"])
    _lebarkan_kolom(ws2, [12, 26, 26, 16, 70])
    if hasil_banding:
        for r, h in enumerate(hasil_banding, start=2):
            ws2.cell(r, 1, h.nrp)
            ws2.cell(r, 2, h.nama_absensi)
            ws2.cell(r, 3, h.nama_gaji)
            ws2.cell(r, 4, h.status)
            ws2.cell(r, 5, "; ".join(h.detail) if h.detail else "")
            fill = OK_FILL if h.status == "COCOK" else BAD_FILL
            for c in range(1, 6):
                ws2.cell(r, c).fill = fill
                ws2.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")

    # Sheet 3: Ringkasan
    ws3 = wb.create_sheet("Ringkasan")
    ws3.cell(1, 1, "Ringkasan Hasil Banding").font = Font(bold=True, size=13)
    if ringkasan_banding:
        baris = [
            ("Total karyawan dibandingkan", ringkasan_banding.get("total")),
            ("Cocok / sinkron", ringkasan_banding.get("cocok")),
            ("Tidak sinkron (perlu diperbaiki)", ringkasan_banding.get("tidak_sinkron")),
            ("Hanya ada di data absensi", ringkasan_banding.get("hanya_di_absensi")),
            ("Hanya ada di struk gaji", ringkasan_banding.get("hanya_di_struk_gaji")),
        ]
        for i, (label, val) in enumerate(baris, start=3):
            ws3.cell(i, 1, label)
            ws3.cell(i, 2, val)
    _lebarkan_kolom(ws3, [36, 12])

    wb.save(path_output)
    return path_output
